import pandas as pd
import openpyxl as openpyxl

class Data:
    def __init__(self):
        self.data_excel = pd.read_excel('data.xlsx')
        self.names_list = list(self.data_excel)
        self.values_list = self.data_excel.values.tolist()

    def get_values(self):

        temp = [str(i[0]) for i in self.values_list]
        self.date = [str(i[:10]) for i in temp]

        self.values = [str(i[1]) for i in self.values_list]

        self.status = [str(i[2]) for i in self.values_list]

        self.type = [str(i[3]) for i in self.values_list]

class Functions(Data):
    def __init__(self):
        super().__init__()
        super().get_values()
        
    def calc(self):
        file1_r = open("bank_status.txt", "r")

        bank_list = file1_r.readlines()

        if len(self.date) -1 > int(bank_list[1]):
            num_1 = bank_list[0]
            num_1 = float(num_1[:-1])
            for i in range(int(bank_list[1])+1, len(self.date)):
                if self.status[i] == "Income":
                    num_1 = num_1 + float(self.values[i])
                    print(num_1)
                else:
                    num_1 = num_1 - float(self.values[i])


            bank_list = [str(round(num_1, 2)) + "\n", str(i)]

            file1_r.close()
            
            file1_w = open("bank_status.txt", "w")
            file1_w.writelines(bank_list)

            file1_w.close()
            print("UPDATED")
        else:
            print("No new entries")


    def calc_date(self):
        date_1 = ""
        num_1 = 0
        for i, v in enumerate(self.date):

            if self.status[i] == "Income":
                value = float(self.values[i])
            else:
                value = -float(self.values[i])

            if v != date_1:
                if date_1 != "": print("| " + date_1 + " | " + str(round(num_1,2)) + " |")
                date_1 = v
                num_1 = value

            else:
                num_1 += value
        print("| " + date_1 + " | " + str(round(num_1, 2)) + " |")
    def insert_data(self):
        plik_excel = openpyxl.load_workbook("data.xlsx")
        arkusz = plik_excel["Arkusz1"]

        def insert_excel(num_row,num_column,cell_value):
            arkusz.cell(row=num_row, column=num_column, value=cell_value)

        number = int(self.names_list[4])

        insert_descision = 1
        while(insert_descision == 1):
            cell_value = input("Podaj date np:(26.06.2023): ")

            insert_excel(number,1,cell_value)
            cell_value = input("Podaj kwote: ")
            insert_excel(number,2,float(cell_value))
            cell_value = input("Podaj status kwoty (Income/Losses): ")
            insert_excel(number,3,cell_value)
            cell_value = input("Podaj typ kowoty (PLN, itp.): ")
            insert_excel(number,4,cell_value)

            yesno = input("Dodać następny rekord?(Tak/Nie): ")
            try:
                if yesno == "Tak":
                    number += 1
                elif yesno == "Nie":
                    number += 1
                    insert_descision = 0
                else:
                    raise Exception
            except:
                print("Blad zla opcja")

        arkusz.cell(row=1, column=5, value=number)
        plik_excel.save("data.xlsx")

#obiekty klas
data_1 = Data()
func = Functions()

#old solutions:
'''
names = data_1.names_list
date = data_1.get_values(0)
values = data_1.get_values(1)
status = data_1.get_values(2)
type = data_1.get_values(3)
'''

print("*** Co chcesz zrobic?: ***")
print("1. Zaaktualizuj status banku")
print("2. Wypisz kwoty na podstawie dat")
print("3. Wprowadz nowe rekordy")
print("4. Wyjdz")

while 1 == 1:
    
    try:
        decision = input("Wybierz jedna z opcji: ")
        if decision == "1":
            func.calc()
            break
        elif decision == "2":
            func.calc_date()
            break
        elif decision == "3":
            func.insert_data()
            break
        elif decision == "4":
            break
        else:
            raise Exception
    except:
        print("Blad zla opcja")
